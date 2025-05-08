from time import sleep

from openpyxl import Workbook

from typing import List

from dune_client.client import DuneClient
from dune_client.query import QueryBase
from dune_client.types import QueryParameter, ParameterType

import pandas as pd

class Dune(DuneClient):
    def __init__(self, api_key, base_url, request_timeout=30000):
        super().__init__(
                api_key=api_key,
                base_url=base_url,
                request_timeout=request_timeout,
            )

    def query(self, query_id, params: List[QueryParameter]):
        query = QueryBase(
            query_id=query_id,
            params=params
        )
        query_result = self.run_query_dataframe(
            query=query
        )
        return query_result

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        pass

api_key = '8WD7smZHIH9MwyvdCLRwJYlryLOBOP80'
base_url = "https://api.dune.com"


TABLE_DICT = {
    "bera": "erc721_berachain.evt_Transfer",
    "arb": "erc721_arbitrum.evt_Transfer",
    "eth": "erc721_ethereum.evt_Transfer",
}

with Dune(api_key=api_key, base_url=base_url) as dune:
    query_id = 4966677

    df = pd.read_excel('input.xlsx')

    result = {}
    sheet3 = [["contract_address_have_no_data"]]

    for index, row in df.iterrows():
        print(f"Hàng {index}: {row.to_dict()}")

        contracts_str = row.to_dict().get('Contract')
        contracts = contracts_str.split("\n")

        chain = row.to_dict().get('Chain').lower()
        point_multiplier = row.to_dict().get('Point muiltiplier ')

        for contract in contracts:
            contract = contract.lower()
            if len(contract) < 10:
                continue
            attempt = 1
            retry_delay = 2
            while True:
                print(f"Attempt {attempt}: {contract}")
                try:
                    params = [
                        QueryParameter("contract_address", ParameterType.TEXT, contract),
                        QueryParameter("table_name", ParameterType.TEXT, TABLE_DICT[chain]),
                    ]

                    dune_data = dune.query(query_id=query_id, params=params)
                    break
                except Exception as e:
                    print(f"Attempt {attempt} got error: {e}")
                    sleep(retry_delay)
                finally:
                    attempt += 1

            number_of_rows = 0
            for index, row in dune_data.iterrows():
                number_of_rows += 1
                d = row.to_dict()
                by_hoder = result.setdefault(d['holder_address'], {})

                by_hoder_address = by_hoder.setdefault(d['contract_address'], {})
                by_hoder_address.setdefault('nft_count', d['nft_count'])
                by_hoder_address.setdefault('point_multiplier', point_multiplier)


            if number_of_rows == 0:
                sheet3.append([contract])
                print("FK_CONTRACT", contract)


    sheet1 = [["hoder_address", "contract_address", "nft_count", "point_multiplier"]]
    sheet2 = [["hoder_address", "m1"]]

    for holder_address, contract_addresses in result.items():
        m1 = 0
        for contract_address, value in contract_addresses.items():
            nft_count = value['nft_count']
            point_multiplier = value['point_multiplier']
            m1 += nft_count * point_multiplier
            sheet1.append([holder_address, contract_address, nft_count, point_multiplier])


        sheet2.append([holder_address, m1])

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "sheet1"
    for row in sheet1:
        ws1.append(row)

    ws2 = wb.create_sheet(title="sheet2")
    for row in sheet2:
        ws2.append(row)

    w3 = wb.create_sheet(title="sheet3")
    for row in sheet3:
        w3.append(row)

    # Lưu file Excel
    wb.save("output_m1.xlsx")